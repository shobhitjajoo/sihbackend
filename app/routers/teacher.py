# app/routers/teacher.py

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List
from datetime import date
from fastapi.responses import StreamingResponse
import io
import openpyxl
from typing import Optional
from app import models, schemas, database
from app.utils.auth_utils import get_current_user

router = APIRouter(prefix="/teacher", tags=["Teacher"])

get_db = database.get_db


# ----------------------------
# ✅ Utility: Ensure teacher scope
# ----------------------------
def get_teacher_user(current_user: dict, db: Session):
    if current_user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Not a teacher")

    teacher = db.query(models.Teacher).filter(models.Teacher.id == current_user["id"]).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    return teacher


# ----------------------------
# 👀 View Students (Read-only)
# ----------------------------
@router.get("/students", response_model=List[schemas.StudentOut])
def get_students(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)
    class_ids = [c.id for c in teacher.classes]
    return db.query(models.Student).filter(models.Student.class_id.in_(class_ids)).all()


# ----------------------------
# 📝 Attendance Management
# ----------------------------

@router.post("/attendance/student/{student_id}")
def mark_student_attendance(
    student_id: int,
    status: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)
    class_ids = [c.id for c in teacher.classes]
    student = (
    db.query(models.Student)
    .filter(models.Student.id == student_id, models.Student.class_id.in_(class_ids))
    .first()
)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in your class")
# Prevent duplicate marking for same date
    existing = (
        db.query(models.Attendance)
        .filter(
            models.Attendance.student_id == student.id,
            models.Attendance.date == date.today(),
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Attendance already marked for today")
    attendance = models.Attendance(
        student_id=student.id,
        teacher_id=teacher.id,
        date=date.today(),
        status=status,
    )
    db.add(attendance)
    db.commit()
    db.refresh(attendance)
    return {"detail": "Attendance marked", "attendance_id": attendance.id}


# ----------------------------
# 📊 Attendance Reports
# ----------------------------
@router.get("/attendance/student/{student_id}")
def get_student_attendance(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)
    class_ids = [c.id for c in teacher.classes]

    student = (
        db.query(models.Student)
        .filter(models.Student.id == student_id, models.Student.class_id.in_(class_ids))
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in your class")

    # Fetch all attendance records for this student
    records = db.query(models.Attendance).filter(
        models.Attendance.student_id == student.id
    ).all()

    total = len(records)
    present = len([r for r in records if r.status.lower() == "present"])
    absent = len([r for r in records if r.status.lower() == "absent"])

    return {
        "student": student.name,
        "total_days": total,
        "present": present,
        "absent": absent,
        "attendance_%": (present / total * 100) if total else 0,
    }


@router.get("/attendance/class")
def get_class_attendance(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)

    result = []

    # Loop through each class the teacher teaches
    for cls in teacher.classes:
        students = db.query(models.Student).filter(models.Student.class_id == cls.id).all()
        stats = []

        for s in students:
            records = db.query(models.Attendance).filter(
                models.Attendance.student_id == s.id
            ).all()

            total = len(records)
            present = len([r for r in records if r.status.lower() == "present"])
            absent = len([r for r in records if r.status.lower() == "absent"])

            stats.append({
                "student": s.name,
                "total_days": total,
                "present": present,
                "absent": absent,
                "attendance_%": (present / total * 100) if total else 0,
            })

        result.append({
            "class_name": cls.name,
            "class_id": cls.id,
            "stats": stats
        })

    return {"teacher_id": teacher.id, "classes": result}

# ----------------------------
# 📤 Export Student Attendance to Excel
# ----------------------------
@router.get("/attendance/student/{student_id}/export")
def export_student_attendance_excel(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)
    class_ids = [c.id for c in teacher.classes]
    student = (
        db.query(models.Student)
        .filter(models.Student.id == student_id, models.Student.class_id.in_(class_ids))
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in your class")

    records = db.query(models.Attendance).filter(models.Attendance.student_id == student.id).all()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append(["Date", "Status"])
    for r in records:
        ws.append([r.date.strftime("%Y-%m-%d"), r.status])

    # Save to bytes
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=attendance_student_{student.id}.xlsx"
        },
    )


# ----------------------------
# 📤 Export Class Attendance to Excel
# ----------------------------
@router.get("/attendance/class/export")
def export_class_attendance_excel(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)
    class_ids = [c.id for c in teacher.classes]
    students = db.query(models.Student).filter(models.Student.class_id.in_(class_ids)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Attendance"

    ws.append(["Student Name", "Date", "Status"])

    for s in students:
        records = db.query(models.Attendance).filter(models.Attendance.student_id == s.id).all()
        for r in records:
            ws.append([s.name, r.date.strftime("%Y-%m-%d"), r.status])

    # Save to bytes
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename=f"attendance_teacher_{teacher.id}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}.xlsx"
        },
    )

# ----------------------------
# 📤 Export Student Attendance to Excel (with filters)
# ----------------------------
@router.get("/attendance/student/{student_id}/export")
def export_student_attendance_excel(
    student_id: int,
    month: Optional[int] = None,   # e.g. 1 = Jan
    year: Optional[int] = None,    # e.g. 2025
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)
    class_ids = [c.id for c in teacher.classes]
    student = (
        db.query(models.Student)
        .filter(models.Student.id == student_id, models.Student.class_id.in_(class_ids))
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in your class")

    query = db.query(models.Attendance).filter(models.Attendance.student_id == student.id)

    # Apply filters
    if month and year:
        query = query.filter(
            db.extract("month", models.Attendance.date) == month,
            db.extract("year", models.Attendance.date) == year,
        )
    if start_date and end_date:
        query = query.filter(models.Attendance.date.between(start_date, end_date))

    records = query.all()

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Date", "Status"])

    for r in records:
        ws.append([r.date.strftime("%Y-%m-%d"), r.status])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=attendance_student_{student.id}.xlsx"
        },
    )


# ----------------------------
# 📤 Export Class Attendance to Excel (with filters)
# ----------------------------
@router.get("/attendance/class/export")
def export_class_attendance_excel(
    month: Optional[int] = None,
    year: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    teacher = get_teacher_user(current_user, db)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Loop through each class
    for cls in teacher.classes:
        ws = wb.create_sheet(title=cls.name)
        ws.append(["Student Name", "Date", "Status"])

        students = db.query(models.Student).filter(models.Student.class_id == cls.id).all()

        for s in students:
            query = db.query(models.Attendance).filter(models.Attendance.student_id == s.id)

            # Apply optional filters
            if month and year:
                query = query.filter(
                    db.extract("month", models.Attendance.date) == month,
                    db.extract("year", models.Attendance.date) == year,
                )
            if start_date and end_date:
                query = query.filter(models.Attendance.date.between(start_date, end_date))

            records = query.all()
            for r in records:
                ws.append([s.name, r.date.strftime("%Y-%m-%d"), r.status])

    # Save workbook to bytes
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"attendance_teacher_{teacher.id}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )