# app/utils/excel_utils.py
import io
from typing import List
from openpyxl import Workbook, load_workbook
from fastapi.responses import StreamingResponse
from app import schemas


# ---------------------------
# EXPORT HELPERS
# ---------------------------
def export_students_to_excel(students: List[dict]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["ID", "Name", "Roll No", "Class"])
    for s in students:
        ws.append([s.get("id"), s.get("name"), s.get("roll_no"), s.get("class_name")])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"},
    )


def export_attendance_to_excel(attendance: List[dict]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Student Name", "Date", "Status"])
    for a in attendance:
        ws.append([a.get("student_name"), a.get("date"), a.get("status")])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance.xlsx"},
    )


# ---------------------------
# IMPORT HELPERS
# ---------------------------
def read_students_excel(file) -> List[schemas.StudentCreate]:
    wb = load_workbook(file)
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        students.append(schemas.StudentCreate(
            name=row[0],
            roll_no=row[1],
            class_id=row[2],
        ))
    return students


def read_teachers_excel(file) -> List[schemas.TeacherCreate]:
    wb = load_workbook(file)
    ws = wb.active
    teachers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        teachers.append(schemas.TeacherCreate(
            name=row[0],
            email=row[1],
            subject=row[2],
            school_id=row[3],
        ))
    return teachers


def read_administrators_excel(file) -> List[schemas.AdministratorCreate]:
    wb = load_workbook(file)
    ws = wb.active
    admins = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        admins.append(schemas.AdministratorCreate(
            name=row[0],
            email=row[1],
            school_id=row[2],
        ))
    return admins

# app/utils/excel_utils.py

def read_classes_excel(file) -> List[schemas.ClassCreate]:
    wb = load_workbook(file)
    ws = wb.active
    classes = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        classes.append(schemas.ClassCreate(
            name=row[0],
            teacher_id=row[1],
        ))
    return classes


def generate_attendance_excel(attendance: List[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["ID", "Student ID", "Teacher ID", "Status", "Date"])
    for a in attendance:
        ws.append([
            a.get("id"),
            a.get("student_id"),
            a.get("teacher_id"),
            a.get("status"),
            a.get("date"),
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream